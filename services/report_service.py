from sqlalchemy.orm import Session
from fastapi import HTTPException
from app.database.models.ProjectDB import Project
from project_management.app.database.models.financial_info_DB import FinancialInfo
from app.database.models.guarantee import Guarantee
from app.database.models.social_security import SocialSecurity
from sqlalchemy import func
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import io
import locale
from typing import Dict, Tuple
from decimal import Decimal, InvalidOperation

# Set locale for number formatting
locale.setlocale(locale.LC_ALL, '')

class ReportService:
    """
    Service class for generating various types of reports in Excel format.
    Handles data validation, report generation, and Excel file creation.
    """
    
    # Report configurations for different report types
    REPORT_CONFIGS = {
        "مشخصات پروژه": {
            "model": Project,
            "columns": [
                "company_id", "company_name", "subject", "project_code", "contract_number",
                "contract_date", "initial_amount", "change_amount", "start_date",
                "duration", "extension_date", "contact_numbers", "archive_number",
                "tax_clearance", "contract_type", "project_status"
            ],
            "labels": [
                "شناسه ملی شرکت", "نام شرکت", "موضوع", "کد پروژه", "شماره قرارداد",
                "تاریخ قرارداد", "مبلغ اولیه قرارداد", "مبلغ افزایش/کاهش", "تاریخ شروع پیمان",
                "مدت پیمان (ماه)", "تمدید مدت پیمان", "شماره تماس", "شماره بایگانی",
                "وضعیت مفاصا حساب", "نوع قرارداد", "وضعیت پروژه"
            ],
            "numeric_fields": ["مبلغ اولیه قرارداد", "مبلغ افزایش/کاهش"]
        },
        "وضعیت مالی": {
            "model": FinancialInfo,
            "columns": [
                "company_id", "record_number", "stage", "request_number", "request_date",
                "invoice_number", "invoice_type", "invoice_amount", "allocation_amount",
                "request_result", "settlement_method", "advance_amortization",
                "partial_amortization", "paid_amount", "remaining_invoice",
                "remaining_advance", "remaining_partial", "contractor_credit",
                "contractor_debit", "total_paid_invoices"
            ],
            "labels": [
                "شناسه ملی شرکت", "شماره ردیف", "مرحله", "شماره نامه درخواست",
                "تاریخ نامه درخواست", "شماره صورت‌وضعیت", "نوع صورت‌وضعیت",
                "مبلغ صورت‌وضعیت", "مبلغ تخصیص", "نتیجه درخواست", "روش تسویه",
                "استهلاک پیش‌پرداخت", "استهلاک علی‌الحساب", "مبلغ پرداخت‌شده",
                "مانده صورت‌وضعیت", "مانده پیش‌پرداخت", "مانده علی‌الحساب",
                "بستانکاری پیمانکار", "بدهکاری پیمانکار", "جمع مبالغ پرداخت‌شده"
            ],
            "numeric_fields": [
                "مبلغ صورت‌وضعیت", "مبلغ تخصیص", "استهلاک پیش‌پرداخت", 
                "استهلاک علی‌الحساب", "مبلغ پرداخت‌شده", "مانده صورت‌وضعیت",
                "مانده پیش‌پرداخت", "مانده علی‌الحساب", "بستانکاری پیمانکار",
                "بدهکاری پیمانکار", "جمع مبالغ پرداخت‌شده"
            ]
        },
        "تضامین": {
            "model": Guarantee,
            "columns": [
                "company_id", "record_number", "guarantee_type", "guarantee_category",
                "guarantee_number", "guarantee_date", "guarantee_amount", "guarantee_bank",
                "guarantee_expiry", "deposit_type", "deposit_date", "deposit_amount",
                "deposit_release_reason", "deposit_released_amount", "deposit_release_date"
            ],
            "labels": [
                "شناسه ملی شرکت", "شماره ردیف", "نوع تضمین", "نوع ضمانت‌نامه",
                "شماره ضمانت‌نامه", "تاریخ ضمانت‌نامه", "مبلغ ضمانت‌نامه", "بانک عامل",
                "تاریخ انقضا", "نوع سپرده", "تاریخ سپرده", "مبلغ سپرده",
                "دلیل آزادسازی", "مبلغ آزادشده", "تاریخ آزادسازی"
            ],
            "numeric_fields": ["مبلغ ضمانت‌نامه", "مبلغ سپرده", "مبلغ آزادشده"]
        },
        "بیمه تأمین اجتماعی": {
            "model": SocialSecurity,
            "columns": ["company_id", "record_number", "insurance_amount", "payment_method"],
            "labels": ["شناسه ملی شرکت", "شماره ردیف", "مبلغ بیمه تأمین اجتماعی", "نحوه پرداخت"],
            "numeric_fields": ["مبلغ بیمه تأمین اجتماعی"]
        }
    }

    @classmethod
    def generate_report(cls, db: Session, company_id: str, report_type: str) -> bytes:
        """
        Generate a report for the specified company and type.
        
        Args:
            db: Database session
            company_id: Company ID to generate report for
            report_type: Type of report to generate
            
        Returns:
            Excel file as bytes
            
        Raises:
            HTTPException: If company_id is invalid or report_type is not supported
        """
        cls._validate_company_id(company_id)
        cls._validate_report_type(report_type)
        
        project = db.query(Project).filter_by(company_id=company_id).first()
        if not project:
            raise HTTPException(
                status_code=404,
                detail=f"شرکت با شناسه ملی {company_id} یافت نشد"
            )

        if report_type == "گزارش کلی":
            return cls._generate_full_report(db, company_id)

        return cls._generate_single_report(db, company_id, report_type)

    @classmethod
    def _validate_company_id(cls, company_id: str) -> None:
        """Validate company ID format"""
        if not company_id:
            raise HTTPException(
                status_code=400,
                detail="شناسه ملی شرکت نمی‌تواند خالی باشد"
            )
        
        if not company_id.isdigit():
            raise HTTPException(
                status_code=400,
                detail="شناسه ملی شرکت باید فقط شامل اعداد باشد"
            )

    @classmethod
    def _validate_report_type(cls, report_type: str) -> None:
        """Validate report type"""
        valid_types = list(cls.REPORT_CONFIGS.keys()) + ["گزارش کلی"]
        if report_type not in valid_types:
            raise HTTPException(
                status_code=400,
                detail=f"نوع گزارش نامعتبر است. انواع معتبر: {', '.join(valid_types)}"
            )

    @classmethod
    def _generate_single_report(cls, db: Session, company_id: str, report_type: str) -> bytes:
        """
        Generate a single report of specified type.
        
        Args:
            db: Database session
            company_id: Company ID
            report_type: Type of report
            
        Returns:
            Excel file as bytes
        """
        config = cls.REPORT_CONFIGS[report_type]
        data = db.query(config["model"]).filter_by(company_id=company_id).all()
        
        if not data:
            raise HTTPException(
                status_code=404,
                detail=f"هیچ داده‌ای برای گزارش {report_type} یافت نشد"
            )

        # Create DataFrame with data
        df = pd.DataFrame([{
            col: getattr(item, col) if getattr(item, col) is not None else ""
            for col in config["columns"]
        } for item in data], columns=config["labels"])

        # Format numeric fields
        cls._format_numeric_fields(df, config["numeric_fields"])

        return cls._create_excel_file(df, report_type)

    @classmethod
    def _generate_full_report(cls, db: Session, company_id: str) -> bytes:
        """
        Generate a comprehensive report including all sections.
        
        Args:
            db: Database session
            company_id: Company ID
            
        Returns:
            Excel file as bytes
        """
        all_data = {}
        totals = {
            "total_paid": Decimal('0'),
            "total_insurance": Decimal('0'),
            "total_deposit": Decimal('0')
        }

        # Collect data for each report type
        for section, config in cls.REPORT_CONFIGS.items():
            data = db.query(config["model"]).filter_by(company_id=company_id).all()
            
            if data:
                df = pd.DataFrame([{
                    col: getattr(item, col) if getattr(item, col) is not None else ""
                    for col in config["columns"]
                } for item in data], columns=config["labels"])
                
                cls._format_numeric_fields(df, config["numeric_fields"])
                all_data[section] = df
                
                # Calculate totals for financial summary
                if section == "وضعیت مالی":
                    totals["total_paid"] = cls._calculate_total(
                        db, FinancialInfo.paid_amount, company_id
                    )
                elif section == "بیمه تأمین اجتماعی":
                    totals["total_insurance"] = cls._calculate_total(
                        db, SocialSecurity.insurance_amount, company_id
                    )
                elif section == "تضامین":
                    totals["total_deposit"] = cls._calculate_total(
                        db, Guarantee.deposit_amount, company_id
                    )

        if not all_data:
            raise HTTPException(
                status_code=404,
                detail="هیچ داده‌ای برای این شرکت یافت نشد"
            )

        return cls._create_full_excel_file(all_data, totals)

    @classmethod
    def _calculate_total(cls, db: Session, column, company_id: str) -> Decimal:
        """Calculate total for a given column"""
        try:
            total = db.query(func.sum(column)).filter_by(company_id=company_id).scalar() or Decimal('0')
            return Decimal(str(total))
        except (ValueError, InvalidOperation):
            return Decimal('0')

    @classmethod
    def _format_numeric_fields(cls, df: pd.DataFrame, numeric_columns: list) -> None:
        """Format numeric fields with locale-aware formatting"""
        for col in df.columns:
            if col in numeric_columns:
                df[col] = df[col].apply(
                    lambda x: locale.format_string("%d", x, grouping=True) 
                    if pd.notnull(x) and str(x).strip() and str(x).strip() != "0" 
                    else "-"
                )

    @classmethod
    def _create_excel_file(cls, df: pd.DataFrame, report_type: str) -> bytes:
        """
        Create an Excel file from DataFrame.
        
        Args:
            df: DataFrame containing report data
            report_type: Title for the report
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.sheet_view.rightToLeft = True
        ws.title = report_type[:31]  # Excel sheet name limit

        # Define styles
        header_font = Font(name="IRANSans", size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A3C34", end_color="1A3C34", fill_type="solid")
        cell_font = Font(name="IRANSans", size=12)
        alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

        # Write data
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.font = cell_font
                cell.alignment = alignment

        # Adjust column widths
        cls._adjust_column_widths(ws)

        # Save to bytes buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @classmethod
    def _create_full_excel_file(cls, all_data: Dict[str, pd.DataFrame], totals: Dict[str, Decimal]) -> bytes:
        """
        Create a comprehensive Excel report with multiple sections.
        
        Args:
            all_data: Dictionary of DataFrames for each section
            totals: Dictionary of total amounts
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.sheet_view.rightToLeft = True
        ws.title = "گزارش کلی"

        # Define styles
        title_font = Font(name="IRANSans", size=16, bold=True, color="1A3C34")
        header_font = Font(name="IRANSans", size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell_font = Font(name="IRANSans", size=12)
        alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

        row_offset = 1
        for section, df in all_data.items():
            # Add section title
            cell = ws.cell(row=row_offset, column=1, value=section)
            cell.font = title_font
            cell.alignment = alignment
            ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=len(df.columns))
            row_offset += 1

            # Add headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=row_offset, column=col_num, value=column_title)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment

            # Add data
            for row_num, row_data in enumerate(df.values, row_offset + 1):
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                    cell.font = cell_font
                    cell.alignment = alignment

            row_offset += len(df) + 3  # Add space between sections

        # Add summary section
        cls._add_summary_section(ws, row_offset, totals, title_font, cell_font, alignment)

        # Adjust column widths
        cls._adjust_column_widths(ws)

        # Save to bytes buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @classmethod
    def _add_summary_section(cls, ws, row_offset: int, totals: Dict[str, Decimal], 
                            title_font: Font, cell_font: Font, alignment: Alignment) -> None:
        """Add summary section to the worksheet"""
        ws.cell(row=row_offset, column=1, value="جمع‌بندی").font = title_font
        ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=2)
        row_offset += 1

        summary_data = [
            ("مجموع پرداختی صورت‌وضعیت‌ها", totals["total_paid"]),
            ("مجموع بیمه تأمین اجتماعی", totals["total_insurance"]),
            ("مجموع سپرده تضامین", totals["total_deposit"])
        ]

        for label, value in summary_data:
            ws.cell(row=row_offset, column=1, value=label).font = cell_font
            ws.cell(row=row_offset, column=2, value=float(value)).font = cell_font
            ws.cell(row=row_offset, column=2).number_format = '#,##0'
            ws.cell(row=row_offset, column=1).alignment = alignment
            ws.cell(row=row_offset, column=2).alignment = alignment
            row_offset += 1

    @classmethod
    def _adjust_column_widths(cls, ws) -> None:
        """Adjust column widths based on content"""
        column_widths = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    col_letter = cell.column_letter
                    cell_len = len(str(cell.value))
                    column_widths[col_letter] = max(column_widths.get(col_letter, 0), cell_len)

        for col_letter, max_length in column_widths.items():
            adjusted_width = min((max_length + 2) * 1.2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width